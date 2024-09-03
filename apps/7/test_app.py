import pytest
import app
import tempfile
import os
import openpyxl

def test_combine_tsv_files_to_excel():
    # Create temporary TSV files
    tsv_files = []
    for i in range(3):
        filename = f"test{i}.tsv"
        with open(filename, "w") as tsv_file:
            tsv_file.write("column1\tcolumn2\n")
            tsv_file.write("data1\tdata2\n")
        tsv_files.append(filename)

    # Run the function
    output_filename = tempfile.mkstemp(suffix=".xlsx")[1]
    app.combine_tsv_files_to_excel(output_filename)

    # Check the output
    workbook = openpyxl.load_workbook(output_filename)
    assert len(workbook.sheetnames) == 4  # Check number of worksheets
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        assert worksheet.max_row == 2  # Check number of rows
        assert worksheet.max_column == 2  # Check number of columns

    # Clean up temporary files
    for filename in tsv_files:
        os.remove(filename)
    os.remove(output_filename)
